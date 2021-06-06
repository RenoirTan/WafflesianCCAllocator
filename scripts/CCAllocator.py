"""
Wafflesian CCA Allocation Algorithm v1.4.0 Beta
===============================================

ATTENTION
---------
Make sure that the input CCA file you input into the program is in the correct format. A template is given /CCA/xlsxfiles/template. You may also read FILEFORMAT.md in /CCA/documentation.
Ensure that all modules in requirements.txt has been installed if you are running the source code file.
Do not distribute without permission.
"""

# If running the application directly, print out entrance message.

if __name__ == "__main__":
    print("Wafflesian CCA Allocation Algorithm v1.3.0 Beta.")

# Import all required modules

import logging
import math
import openpyxl
import os
import pathlib
import random
import traceback

# Important variables
ALOB = None  # Stores the object for class Allocation
TPOB = None  # Stores the object for class Template
PATH = ""
FILENAME = ""
CCALIST = []
CCAALIASES = {}
CCATYPE = {}
SHEETORDER = {
    "studentList": None,
    "healthStats": None,
    "music": None,
    "art": None,
    "special": None,
    "CCAList": None,
    "choices": None,
}
CHOICES = {"main": 1, "other": None}

# Setup error logging

logging.basicConfig(
    level=logging.DEBUG,
    filename="log.log",
    filemode="a",
    format="%(process)d - %(asctime)s - %(levelname)s - %(message)s",
    datefmt="%d-%b-%y %H:%M:%S",
)

if __name__ == "__main__":
    logging.info("CCAllocator: Program run as main.")
else:
    logging.info("CCAllocator: Program run as module.")

# Error Classes


class CCAllocatorError(Exception):
    """Base error for everything in this program."""

    pass


class UnknownError(CCAllocatorError):
    """Error occurring in methods but from previously unknown sources."""

    pass


class MissingSheetError(CCAllocatorError):
    """Raise this error if vital sheets are missing."""

    pass


class MissingDataError(CCAllocatorError):
    """Raise this error if vital pieces of information in sheets are not found."""

    pass


class MissingParameterError(CCAllocatorError):
    """Raise this error if not enough arguments are passed through the functions or methods of this program."""

    pass


class ErroneousDataError(CCAllocatorError):
    """Raise this error if erroneous data is encountered during execution."""

    pass


class ErroneousParameterError(CCAllocatorError):
    """Raise this error if erroneous arguments are passed into any function or methods in this program."""

    pass


class UnableToOpenFileError(CCAllocatorError):
    """Raise this error if input file cannot be accessed."""

    pass


class UnableToSaveFileError(CCAllocatorError):
    """Raise this error if input file cannot be saved."""

    pass


class WrongFileTypeError(ErroneousParameterError):
    """Raise this error if a file that does not have a .xlsx extension is used to try and open the file."""

    pass


# Allocation Class


class Allocation:
    """
    Class for methods and attributes for allocating students to CCAs.

    ATTRIBUTES
    ----------
        self.Path -- Path to file
        self.FileName -- File name
        self.CCAList -- List of CCAs
        self.CCAALiases -- Dictionary of CCA Aliases
        self.CCAType -- Dictionary of CCA type
        self.SheetOrder -- Dictionary of the order of administrative sheets
        self.Workbook -- Object to opened input file
        self.Worksheets -- Dictionary of objects of sheets
        self.OutputWorkbook -- Object to opened output file
        self.OutputStudentList -- Object to student list of output file
        self.StudentData -- List of student data
        self.CCAData -- List of CCA data
    """

    def __init__(self, **kwargs):
        """
        Sets up attributes for later methods.

        **kwargs Parameters
        -------------------
        path: str
            Path to file, must include final slash.
            Example: 'C:\\Path\\'
        fileName: str
            Name of file, must be a .xlsx file.
            Example: 'input_file.'
        listOfCCAs: list
            List of CCAs, program will assume that these are all the CCAs in the input file and that they are the CCAs' standard names, otherwise, that CCA will be ignored.
            Example: ['Infocomm','Swimming']

        Other Parameters
        ----------------
        CCAAliases: dict of list
            Dictionary of lists of CCA aliases. Standard CCA name in keys and aliases in value. Aliases are variants of the CCA's standard name which the program might bump into (inclusive of upper/lower case variants). REQUIRES CCAList
            Example: {'Infocomm':['Computer','Coding'],'Swimming':['swimming']}
        CCAType: dict of str
            Dictionary of string of type of CCA. By default, each CCA will be treated as 'b' or basic. However, you can indicate that the CCA is music (m) or art (a). REQUIRES CCAList
            Example: {'Guitar Ensemble':'m'}
        sheetOrder: dict of int or None
            Dictionary of positional index of administrative sheets as they appear in the input file, with the first sheet being 1 (instead of 0). If the admin sheet is not in the input file, indicate None. Accepted admin sheets: studentList, healthStats, music, art, special, CCAList, choices. REQUIRES CCAList
            Example: {'studentList':1,'choices':2,'CCAList':3}
        choices: dict of int or None
            Dictionary of how many choices each student can have, divided into 'main' and 'other'.
            Example: {'main':9,'other':None}

        Raises
        ------
        TypeError
            If parameters entered in are not of the correct type.
        MissingParameterError
            If compulsory parameters are not entered in.
        WrongFileTypeError
            If the input file type is not .xlsx.
        """
        print("New CCAllocator.Allocation object created.")
        logging.info(
            "Allocation.__init__: Setting up new CCAllocator.Allocation object."
        )
        ### Checks whether all the inputs are all the correct data type to prevent errors being thrown later hopefully. ###
        self.Path = ""  # Path to file
        self.FileName = ""  # File name
        self.CCAList = []  # List of CCAs (optional, conditional)
        self.CCAAliases = {}  # Dictionary of CCAAliases (optional)
        self.CCAType = {}  # Dictionary of CCA types (optional)
        self.SheetOrder = {}  # Order of sheets
        self.Choices = {
            "main": 1,
            "other": None,
        }  # How many CCA choices each student can make
        self.Workbook = None  # Object for the input workbook
        self.Worksheets = {}  # Dictionary of objects of sheets
        self.OutputWorkbook = None  # Object for the output workbook
        self.OutputStudentList = None  # Object for the output workbook`s student list
        self.StudentData = []  # List of student data
        self.CCAData = []  # List of CCA data
        self.HealthStatsData = {}  # Dictionary of general healthStats data
        self.Allocated = False  # Boolean for whether students have been allocated.
        if "path" not in kwargs.keys():
            logging.error("Allocation.__init__: Could not find path in keys.")
            raise MissingParameterError(
                "Path is required for this method to run. Input as such: Allocation.__init__(...,path=`C:/ExampleFolder/ExampleDirectory/`,...)."
            )
        else:
            try:
                assert type(kwargs["path"]) == str, "Wrong data type."
            except AssertionError:
                logging.error("Allocation.__init__: Path is not a string.")
                raise TypeError(
                    "Path must be a string. Type of path:", type(kwargs["path"])
                )
            else:
                self.Path = kwargs["path"]
        if "fileName" not in kwargs.keys():
            logging.error("Allocation.__init__: Could not find fileName in keys.")
            raise MissingParameterError(
                "FileName is required for this method to run. Input as such: Allocation.__init__(...,fileName=`CCAfile.xlsx`,...)."
            )
        else:
            try:
                assert type(kwargs["fileName"]) == str, "Wrong data type."
            except AssertionError:
                logging.error("Allocation.__init__: fileName is not a string.")
                raise TypeError(
                    "FileName must be a string. Type of fileName:",
                    type(kwargs["fileName"]),
                )
            else:
                if kwargs["fileName"].endswith(".xlsx"):
                    self.FileName = kwargs["fileName"]
                else:
                    logging.error("Allocation.__init__: Bad file type.")
                    raise WrongFileTypeError(
                        "Wrong file type. Make sure it is a .xlsx file."
                    )
        if "listOfCCAs" in kwargs.keys():
            try:
                assert type(kwargs["listOfCCAs"]) == list, "Wrong data type."
            except AssertionError:
                logging.error("Allocation.__init__: listOfCCAs is not a list.")
                raise TypeError(
                    "ListOfCCAs must be a list. Type of listOfCCAs:",
                    type(kwargs["listOfCCAs"]),
                )
            else:
                self.CCAList = kwargs["listOfCCAs"]
        else:
            logging.error("Allocation.__init__: listOfCCAs required.")
            raise MissingParameterError(
                "listOfCCAs is required to ensure that all CCAs get included during allocation."
            )
        if "CCAAliases" in kwargs.keys():
            if "listOfCCAs" not in kwargs.keys():
                logging.error(
                    "Allocation.__init__: CCAAliases called but listOfCCAs not in kwargs."
                )
                raise MissingParameterError(
                    "CCAAliases requires you to have listOfCCAs as well in order to prevent errors from cropping up."
                )
            else:
                try:
                    assert type(kwargs["CCAAliases"]) == dict, "Wrong data type."
                except AssertionError:
                    logging.error(
                        "Allocation.__init__: CCAAliases is not a dictionary."
                    )
                    raise TypeError(
                        "CCAAliases must be a dictionary. Type of CCAAliases:",
                        type(kwargs["CCAAliases"]),
                    )
                else:
                    self.CCAAliases = kwargs["CCAAliases"]
        if "CCAType" in kwargs.keys():
            if "listOfCCAs" not in kwargs.keys():
                logging.error(
                    "Allocation__init__: CCAType called but listOfCCAs not in kwargs."
                )
                raise MissingParameterError(
                    "CCAType requires you to have listOfCCAs as well in order to prevent errors from cropping up."
                )
            else:
                try:
                    assert type(kwargs["CCAType"]) == dict, "Wrong data type."
                except AssertionError:
                    logging.error("Allocation.__init__: CCAType not a dictionary.")
                    raise TypeError(
                        "CCAType must be a dictionary. Type of CCAType:",
                        type(kwargs["CCAType"]),
                    )
                else:
                    self.CCAType = kwargs["CCAType"]
        if "sheetOrder" in kwargs.keys():
            try:
                assert type(kwargs["sheetOrder"] == dict), "Wrong data type."
            except AssertionError:
                logging.error("Allocation.__init__: sheetOrder is not a dictionary.")
                raise TypeError(
                    "SheetOrder must be a dictionary. Type of sheetOrder:",
                    type(kwargs["CCAAliases"]),
                )
            else:
                self.SheetOrder = kwargs["sheetOrder"]
        if "numberOfChoices" in kwargs.keys():
            try:
                assert type(kwargs["numberOfChoices"] == dict), "Wrong data type."
            except AssertionError:
                logging.error(
                    "Allocation.__init__: numberOfChoices is not a dictionary."
                )
                raise TypeError(
                    "NumberOfChoices must be a dictionary. Type of numberOfChoices:",
                    type(kwargs["numberOfChoices"]),
                )
            else:
                if "main" in kwargs["numberOfChoices"].keys():
                    if isinstance(kwargs["numberOfChoices"]["main"], int):
                        self.Choices["main"] = kwargs["numberOfChoices"]["main"]
                        logging.info(
                            "Allocation.__init__: numberOfChoices/main:"
                            + str(kwargs["numberOfChoices"]["main"])
                        )
                    else:
                        logging.info(
                            "Allocation.__init__: numberOfChoices/main not int, assuming 1."
                        )
                else:
                    logging.info("Allocation.__init__: numberOfChoices missing main.")
                    print(
                        "numberOfChoices does not have `main` as one of its keys, method will assume that main is 1."
                    )
                if "other" in kwargs["numberOfChoices"].keys():
                    if isinstance(kwargs["numberOfChoices"]["other"], int):
                        self.Choices["other"] = kwargs["numberOfChoices"]["other"]
                        logging.info(
                            "Allocation.__init__: numberOfChoices/other:"
                            + str(kwargs["numberOfChoices"]["other"])
                        )
        print("Object successfully created.")
        logging.info("Allocation.__init__: Setup complete.")

    def OpenFile(self):
        """Opens file specified in Allocation.__init__

        Raises
        ------
        ErroneousParameterError
            If the input file specified in Allocation.__init__ could not be found.
        MissingSheetError
            If there are no sheets in the input file.
        UnknownError
            If an uncaught error was raised. Treated as critical.
        """
        print("Opening input CCA file.")
        print("File path: %s" % pathlib.Path(self.Path, self.FileName))
        logging.info("Allocation.OpenFile: Opening input CCA file.")
        try:
            self.Workbook = openpyxl.load_workbook(
                pathlib.Path(self.Path, self.FileName)
            )
        except FileNotFoundError:
            logging.error(
                "Allocation.OpenFile: Could not find file with path and file name provided."
            )
            raise ErroneousParameterError(
                "Could not find the file specified. Please modify self.Path or self.FileName or call Allocation.__init__ again. File path: "
                + self.Path
                + self.FileName
            )
        except Exception as e:
            logging.critical(
                "Allocation.OpenFile: Unknown error occurred when opening file.",
                exc_info=True,
            )
            raise UnknownError(
                "Unable to open input file. Please send error.log file to the creators.",
                e,
            )
        if len(self.Workbook.get_sheet_names()) < 1:
            logging.error(
                "Allocation.OpenFile: No files located in input file provided."
            )
            raise MissingSheetError("No files located in input file provided.")
        else:
            try:
                for sheetName in self.Workbook.get_sheet_names():
                    self.Worksheets[sheetName] = self.Workbook.get_sheet_by_name(
                        sheetName
                    )
            except Exception as e:
                logging.critical(
                    "Allocation.OpenFile: Critical error occurred when getting sheet objects from input file.",
                    exc_info=True,
                )
                raise UnknownError(
                    "Unidentifiable error occurred when opening worksheets in input file. Please send error.log to creators.",
                    e,
                )
        print("File successfully opened.")
        logging.info("Allocation.OpenFile: Input file opened.")

    def Setup(self):
        """Configures attributes to be compatible with the input file.

        Raises
        ------
        UnknownError
            If an uncaught error was raised. Treated as critical error.
        """
        print("Setting up attributes.")
        logging.info("Allocation.Setup: Setting up attributes.")
        # Configure CCA aliases
        tempdict = self.CCAAliases.copy()  # Make a temporary dictionary for CCAAliases
        self.CCAAliases = {}  # Clear CCAAliases
        # Checks to see if CCA has name in CCAList before reconfiguring CCAAliases to have the aliases be keys for easier retrieval of data
        try:
            # Swap keys and values in CCAAliases
            for CCAStandardName in tempdict.keys():
                if CCAStandardName in self.CCAList:
                    self.CCAAliases[CCAStandardName] = CCAStandardName
                    for aliases in tempdict[CCAStandardName]:
                        self.CCAAliases[aliases] = CCAStandardName
            # Adds remaining CCAs to CCAAliases
            for CCAName in self.CCAList:
                if CCAName not in self.CCAAliases.values():
                    self.CCAAliases[CCAName] = CCAName
        except Exception as e:
            logging.critical(
                "Allocation.Setup: Unknown error occurred when configuring CCAAliases.",
                exc_info=True,
            )
            raise UnknownError(e)
        # Configure SheetOrder and Worksheets
        try:
            templist = [
                "studentList",
                "healthStats",
                "music",
                "art",
                "special",
                "CCAList",
                "choices",
            ]
            for adminSheet, position in self.SheetOrder.items():
                if isinstance(position, int):
                    self.SheetOrder[adminSheet] = list(self.Worksheets.keys())[
                        position - 1
                    ]
                else:
                    self.SheetOrder[adminSheet] = None
            for adminSheet in templist:
                if adminSheet not in self.SheetOrder.keys():
                    self.SheetOrder[adminSheet] = None
        except Exception as e:
            logging.critical(
                "Allocation.Setup: Unknown error occurred when configuring SheetOrder.",
                exc_info=True,
            )
            raise UnknownError(e)
        # Reconfigure Worksheets names
        try:
            # Change sheet names to standard names
            for CCASheetName in self.Worksheets.keys():
                for CCAStandardName in self.CCAAliases.keys():
                    if CCASheetName == CCAStandardName:
                        self.Worksheets[
                            self.CCAAliases[CCAStandardName]
                        ] = self.Worksheets.pop(CCASheetName)
        except Exception as e:
            logging.critical(
                "Allocation.Setup: Unknown error occurred when configuring worksheet names.",
                exc_info=True,
            )
            raise UnknownError(e)
        # Assign CCA types
        try:
            for CCA in self.CCAList:
                if CCA not in list(self.CCAType.keys()):
                    self.CCAType[CCA] = "b"
        except Exception as e:
            logging.critical(
                "Allocation.Setup: Unknown error occurred when configuring CCA Types.",
                exc_info=True,
            )
            raise UnknownError(e)
        # Add in the name of CCAs who do not have sheets for them
        try:
            for CCA in self.CCAList:
                if CCA not in self.Worksheets.keys():
                    self.Worksheets[CCA] = None
        except Exception as e:
            logging.critical(
                "Allocation.Setup: UnknownError occurred when adding CCA names.",
                exc_info=True,
            )
            raise UnknownError(e)
        print("Setup complete.")
        # print(self.Path);print(self.FileName);print(self.CCAList);print(self.CCAAliases);print(self.CCAType);print(self.SheetOrder);print(self.Workbook);print(self.Worksheets) # Test print
        logging.info("Allocation.Setup: Setup complete.")

    def GetData(self):
        """Obtains all required data from input file.

        Raises
        ------
        MissingSheetError
            If vital admin sheets are not found.
        MissingDataError
            If vital data was not found in certain sheets.
        ErroneousDataError
            If data in the input file does not match the parameters indicated in Allocation.__init__
        ValueError
            If uncaught ValueError raised from lower levels of error catching.
        UnknownError
            If an unknown error was raised. Treated as a critical error.
        """
        print("Retrieving data.")
        logging.info("Allocation.GetData: Retrieving data.")
        if self.SheetOrder["studentList"] == None:
            logging.error(
                "Allocation.GetData: Sheet position of studentList not found."
            )
            raise MissingSheetError(
                "The sheet position assigned to studentList is none. Please modify self.SheetOrder or redo Allocation.__init__ and try again."
            )
        if self.SheetOrder["healthStats"] == None:
            logging.info("Allocation.GetData: No healthStats")
        else:
            logging.info("Allocation.GetData: healthStats present")
            if self.Worksheets[self.SheetOrder["healthStats"]].max_column >= 6:
                logging.info("Allocation.GetData: Gathering healthStats parameters")
                for healthStatsParameter in range(
                    6, self.Worksheets[self.SheetOrder["healthStats"]].max_column + 1
                ):
                    self.HealthStatsData[
                        self.Worksheets[self.SheetOrder["healthStats"]]
                        .cell(row=1, column=healthStatsParameter)
                        .value
                    ] = {"sum": 0, "validValues": 0, "average": 0}
            else:
                logging.info("Allocation.GetData: No parameters in healthStats")
        if self.SheetOrder["music"] == None:
            logging.info("Allocation.GetData: No music")
        if self.SheetOrder["art"] == None:
            logging.info("Allocation.GetData: No art")
        if self.SheetOrder["special"] == None:
            logging.info("Allocation.GetData: No special")
        if self.SheetOrder["choices"] == None:
            logging.error("Allocation.GetData: No choices.")
            raise MissingSheetError(
                "The sheet position assigned to choices is none. Please modify self.SheetOrder or redo Allocation.__init__ and try again."
            )
        if self.SheetOrder["CCAList"] == None:
            logging.error("Allocation.GetData: No CCAList")
            raise MissingSheetError(
                "The sheet position assigned to CCAList is none. Please modify self.SheetOrder or redo Allocation.__init__ and try again."
            )
        if self.Worksheets[self.SheetOrder["studentList"]].max_row < 2:
            logging.error("Allocation.GetData: Max row of student list is less than 1.")
            raise MissingSheetError(
                "No students found in Student List. Please add some."
            )
        else:
            logging.info("Allocation.GetData: Retrieving data for students.")
            print("Retrieving data for students.")
            for student in range(
                2, self.Worksheets[self.SheetOrder["studentList"]].max_row + 1
            ):
                tempdict = {}  # Temporary dictionary for each student
                # Student list
                tempdict["no"] = (
                    self.Worksheets[self.SheetOrder["studentList"]]
                    .cell(row=student, column=1)
                    .value
                )
                tempdict["name"] = (
                    self.Worksheets[self.SheetOrder["studentList"]]
                    .cell(row=student, column=3)
                    .value
                )
                tempdict["id"] = (
                    self.Worksheets[self.SheetOrder["studentList"]]
                    .cell(row=student, column=4)
                    .value
                )
                tempdict["class"] = (
                    self.Worksheets[self.SheetOrder["studentList"]]
                    .cell(row=student, column=5)
                    .value
                )
                # healthStats
                # If there is a healthStats worksheet,
                #   For all students in the worksheet,
                #       If the name or id matches the student currently retrieving data for i.e. student i
                #           If healthStats data exists, retrieve data from col 6 onwards with key as header and value as student`s value
                if self.SheetOrder["healthStats"] != None:
                    tempdict["healthStats"] = {}
                    for rowOfStudentInAdminSheet in range(
                        2, self.Worksheets[self.SheetOrder["healthStats"]].max_row + 1
                    ):
                        if (
                            self.Worksheets[self.SheetOrder["healthStats"]]
                            .cell(row=rowOfStudentInAdminSheet, column=2)
                            .value
                            == tempdict["name"]
                            or self.Worksheets[self.SheetOrder["healthStats"]]
                            .cell(row=rowOfStudentInAdminSheet, column=3)
                            .value
                            == tempdict["id"]
                        ):
                            try:
                                if (
                                    self.Worksheets[
                                        self.SheetOrder["healthStats"]
                                    ].max_column
                                    >= 6
                                ):
                                    for healthStatsParameter in range(
                                        6,
                                        self.Worksheets[
                                            self.SheetOrder["healthStats"]
                                        ].max_column
                                        + 1,
                                    ):
                                        tempdict["healthStats"][
                                            self.Worksheets[
                                                self.SheetOrder["healthStats"]
                                            ]
                                            .cell(row=1, column=healthStatsParameter)
                                            .value
                                        ] = (
                                            self.Worksheets[
                                                self.SheetOrder["healthStats"]
                                            ]
                                            .cell(
                                                row=rowOfStudentInAdminSheet,
                                                column=healthStatsParameter,
                                            )
                                            .value
                                        )
                                        # Get average of each health parameter by adding each students value of each parameter into the parameter's sum and valid student values count into the self.HealthStatsData of the parameter and calculating the average simultaneously to fit everything in one if block
                                        if isinstance(
                                            self.Worksheets[
                                                self.SheetOrder["healthStats"]
                                            ]
                                            .cell(
                                                row=rowOfStudentInAdminSheet,
                                                column=healthStatsParameter,
                                            )
                                            .value,
                                            (int, float),
                                        ):
                                            self.HealthStatsData[
                                                self.Worksheets[
                                                    self.SheetOrder["healthStats"]
                                                ]
                                                .cell(
                                                    row=1, column=healthStatsParameter
                                                )
                                                .value
                                            ]["sum"] += (
                                                self.Worksheets[
                                                    self.SheetOrder["healthStats"]
                                                ]
                                                .cell(
                                                    row=rowOfStudentInAdminSheet,
                                                    column=healthStatsParameter,
                                                )
                                                .value
                                            )
                                            self.HealthStatsData[
                                                self.Worksheets[
                                                    self.SheetOrder["healthStats"]
                                                ]
                                                .cell(
                                                    row=1, column=healthStatsParameter
                                                )
                                                .value
                                            ]["validValues"] += 1
                                            if (
                                                self.HealthStatsData[
                                                    self.Worksheets[
                                                        self.SheetOrder["healthStats"]
                                                    ]
                                                    .cell(
                                                        row=1,
                                                        column=healthStatsParameter,
                                                    )
                                                    .value
                                                ]["validValues"]
                                                != 0
                                            ):
                                                self.HealthStatsData[
                                                    self.Worksheets[
                                                        self.SheetOrder["healthStats"]
                                                    ]
                                                    .cell(
                                                        row=1,
                                                        column=healthStatsParameter,
                                                    )
                                                    .value
                                                ]["average"] = (
                                                    self.HealthStatsData[
                                                        self.Worksheets[
                                                            self.SheetOrder[
                                                                "healthStats"
                                                            ]
                                                        ]
                                                        .cell(
                                                            row=1,
                                                            column=healthStatsParameter,
                                                        )
                                                        .value
                                                    ]["sum"]
                                                    / self.HealthStatsData[
                                                        self.Worksheets[
                                                            self.SheetOrder[
                                                                "healthStats"
                                                            ]
                                                        ]
                                                        .cell(
                                                            row=1,
                                                            column=healthStatsParameter,
                                                        )
                                                        .value
                                                    ]["validValues"]
                                                )
                            except Exception as e:
                                logging.critical(
                                    "Allocation.GetData: An unknown error occurred when retrieving healthStats data.",
                                    exc_info=True,
                                )
                                raise UnknownError(
                                    "An unknown error occurred when retieving healthStats. Row:",
                                    rowOfStudentInAdminSheet,
                                    "Column:",
                                    healthStatsParameter,
                                    "Error statement:",
                                )
                # Music, art: find students who are interested in said subjects from both sheets if they exist
                # Music
                tempdict["musicInterest"] = False
                tempdict["musicRemarks"] = None
                if self.SheetOrder["music"] != None:
                    for rowOfStudentInAchievementSheet in range(
                        2, self.Worksheets[self.SheetOrder["music"]].max_row + 1
                    ):
                        if (
                            self.Worksheets[self.SheetOrder["music"]]
                            .cell(row=rowOfStudentInAchievementSheet, column=1)
                            .value
                            == tempdict["id"]
                        ):
                            tempdict["musicInterest"] = True
                # Art
                tempdict["artInterest"] = False
                tempdict["artRemarks"] = None
                if self.SheetOrder["art"] != None:
                    for rowOfStudentInAchievementSheet in range(
                        2, self.Worksheets[self.SheetOrder["art"]].max_row + 1
                    ):
                        if (
                            self.Worksheets[self.SheetOrder["art"]]
                            .cell(row=rowOfStudentInAchievementSheet, column=1)
                            .value
                            == tempdict["id"]
                        ):
                            tempdict["artInterest"] = True
                # Get students achievements
                # Special
                tempdict["special"] = False
                tempdict["domain"] = None
                tempdict["specialCCA"] = None
                if self.SheetOrder["special"] != None:
                    for rowOfStudentInAchievementSheet in range(
                        2, self.Worksheets[self.SheetOrder["special"]].max_row + 1
                    ):
                        if (
                            self.Worksheets[self.SheetOrder["special"]]
                            .cell(row=rowOfStudentInAchievementSheet, column=1)
                            .value
                            == tempdict["name"]
                        ):
                            tempdict["special"] = True
                            tempdict["domain"] = (
                                self.Worksheets[self.SheetOrder["special"]]
                                .cell(row=rowOfStudentInAchievementSheet, column=2)
                                .value
                            )
                            tempdict["specialCCA"] = (
                                self.Worksheets[self.SheetOrder["special"]]
                                .cell(row=rowOfStudentInAchievementSheet, column=3)
                                .value
                            )
                # Obtain student's choices
                tempdict["mainChoices"] = []
                tempdict["otherChoices"] = []
                if self.SheetOrder["choices"] != None:
                    for rowOfStudentInAchievementSheet in range(
                        2, self.Worksheets[self.SheetOrder["choices"]].max_row + 1
                    ):
                        if (
                            self.Worksheets[self.SheetOrder["choices"]]
                            .cell(row=rowOfStudentInAchievementSheet, column=2)
                            .value
                            == tempdict["id"]
                        ):
                            for columnOfChoices in range(6, self.Choices["main"] + 6):
                                try:
                                    tempdict["mainChoices"].append(
                                        self.CCAAliases[
                                            self.Worksheets[self.SheetOrder["choices"]]
                                            .cell(
                                                row=rowOfStudentInAchievementSheet,
                                                column=columnOfChoices,
                                            )
                                            .value
                                        ]
                                    )
                                except KeyError as e:
                                    logging.error(
                                        "Allocation.GetData: While searching for main choices, could not get standard name for this CCA:"
                                        + self.Worksheets[self.SheetOrder["choices"]]
                                        .cell(
                                            row=rowOfStudentInAchievementSheet,
                                            column=columnOfChoices,
                                        )
                                        .value
                                    )
                                    raise ErroneousDataError(
                                        "One of the CCA choices indicated in",
                                        self.SheetOrder["choices"],
                                        "could not be found in CCA choices or aliases. Please make sure that the info is correct. Location of error in sheet - Row:",
                                        rowOfStudentInAchievementSheet,
                                        "Column:",
                                        columnOfChoices,
                                    )
                                except Exception as e:
                                    logging.critical(
                                        "Allocation.GetData: Unknown error occurred when retrieving main choices.",
                                        exc_info=True,
                                    )
                                    raise UnknownError(
                                        "An unknown error occurred when retrieving main choices. Row:",
                                        rowOfStudentInAchievementSheet,
                                        "Column:",
                                        columnOfChoices,
                                        "Error statement:",
                                        e,
                                    )
                            if self.Choices["other"] != None:
                                for columnOfChoices in range(
                                    self.Choices["main"] + 6,
                                    self.Choices["main"] + self.Choices["other"] + 6,
                                ):
                                    tempdict["otherChoices"].append(
                                        self.Worksheets[self.SheetOrder["choices"]]
                                        .cell(
                                            row=rowOfStudentInAchievementSheet,
                                            column=columnOfChoices,
                                        )
                                        .value
                                    )
                tempdict["shortlist"] = {}
                tempdict["CCA"] = None
                tempdict["CCARank"] = None
                tempdict["choiceNumbers"] = {"main": 0, "other": None}
                tempdict["fitnessScore"] = 0
                tempdict["locationOfAllocation"] = None
                tempdict["suitability"] = {}
                self.StudentData.append(tempdict)
        logging.info("Allocation.GetData: Retrieving data for CCAs.")
        print("Retrieving data for CCAs.")
        for CCA in range(2, self.Worksheets[self.SheetOrder["CCAList"]].max_row + 1):
            tempdict = {}
            if (
                self.Worksheets[self.SheetOrder["CCAList"]]
                .cell(row=CCA, column=2)
                .value
                in self.CCAAliases.values()
            ):
                tempdict["name"] = self.CCAAliases[
                    self.Worksheets[self.SheetOrder["CCAList"]]
                    .cell(row=CCA, column=2)
                    .value
                ]
                # print(tempdict['name'])
                tempdict["allocated"] = (
                    self.Worksheets[self.SheetOrder["CCAList"]]
                    .cell(row=CCA, column=4)
                    .value
                )
                tempdict["quota"] = (
                    self.Worksheets[self.SheetOrder["CCAList"]]
                    .cell(row=CCA, column=5)
                    .value
                )
                tempdict["shortlist"] = []
                tempdict["selectedStudents"] = []
                tempdict["Popularity"] = 0
                # Get shortlist data if CCA has its shortlist worksheet, then add shortlist
                if (
                    self.Worksheets[tempdict["name"]] != None
                    and self.Worksheets[tempdict["name"]].max_row > 1
                ):
                    for studentInShortlist in range(
                        2, self.Worksheets[tempdict["name"]].max_row + 1
                    ):
                        tempdict["shortlist"].append(
                            self.Worksheets[tempdict["name"]]
                            .cell(row=studentInShortlist, column=2)
                            .value
                        )
                self.CCAData.append(tempdict)
        # Check total students in CCA and actual number of students
        print("Validating data.")
        logging.info("Allocation.GetData: Checking number of students.")
        tempint = 0  # Temporary integer
        for CCA in self.CCAData:
            try:
                tempint += CCA["allocated"]
            except TypeError:
                logging.error(
                    "Allocation.GetData: Allocatable number of students in CCA is not an integer."
                )
                raise ErroneousDataError(
                    "Number of students allocated to CCA must be an integer. CCA: "
                    + CCA["name"]
                    + " Value: "
                    + CCA["allocated"]
                )
        if tempint != len(self.StudentData):
            logging.error("Allocation.GetData: Allocated-Students discrepancy.")
            raise ErroneousDataError(
                "Number of students and total number of allocated students in CCA does not match. Total number of students: "
                + str(len(self.StudentData))
                + " Total number of students all CCAs can have: "
                + str(tempint)
            )
        print(
            "Total number of students: "
            + str(len(self.StudentData))
            + " Total number of students all CCAs can have: "
            + str(tempint)
        )
        # Extract each student's shortlist from each CCA
        logging.info(
            "Allocation.GetData: Matching shortlists of each CCA to each student."
        )
        print("Matching shortlists of each CCA to each student.")
        for student in range(len(self.StudentData)):
            for CCA in range(len(self.CCAData)):
                try:
                    if (
                        self.StudentData[student]["name"]
                        in self.CCAData[CCA]["shortlist"]
                    ):
                        self.StudentData[student]["shortlist"][
                            self.CCAData[CCA]["name"]
                        ] = self.CCAData[CCA]["shortlist"].index(
                            self.StudentData[student]["name"]
                        )
                except ValueError as e:
                    logging.critical(
                        "Allocation.GetData: ValueError when assigning shortlists.",
                        exc_info=True,
                    )
                    raise ValueError(e)
        # Assign student's fitnessScore from healthStats if it exists.
        if self.SheetOrder["healthStats"] != None and self.HealthStatsData != {}:
            for student in range(len(self.StudentData)):
                tempint = 0
                for (
                    healthStatsParameter,
                    healthStatsParameterValue,
                ) in self.HealthStatsData.items():
                    try:
                        if (
                            self.StudentData[student]["healthStats"][
                                healthStatsParameter
                            ]
                            != None
                        ):
                            if healthStatsParameter == "weight":
                                tempint += (
                                    healthStatsParameterValue["average"]
                                    / self.StudentData[student]["healthStats"][
                                        healthStatsParameter
                                    ]
                                )
                            else:
                                tempint += (
                                    self.StudentData[student]["healthStats"][
                                        healthStatsParameter
                                    ]
                                    / healthStatsParameterValue["average"]
                                )
                    except KeyError:
                        logging.warning(
                            "Allocation.GetData: Key error when assigning student fitnessScore. Student: "
                            + str(student + 1)
                            + " healthStatsParameter: "
                            + healthStatsParameter
                            + " value: "
                            + str(self.StudentData[student]["healthStats"])
                        )
                self.StudentData[student]["fitnessScore"] = tempint / len(
                    self.HealthStatsData
                )
        # Initiate suitability
        for student in range(len(self.StudentData)):
            for CCA in self.CCAData:
                self.StudentData[student]["suitability"][CCA["name"]] = 0
        logging.info("Allocation.GetData: Data successfully retrieved.")
        print("Data retrieved.")

    def Allocate(self):
        """Allocates students based on data received.

        Raises
        ------
        Exception
            If an unknown error occurs
        """

        # You will notice the keyword location multiple times in this method.
        # What does it mean?
        # It tells us where (and thus how) the student was allocated.
        # For example, when allocating by CCA shortlist and student choice,
        # the student allocated there would be allocated at location 0 (because 0 is the 1st number in Python)

        def AddStudent(
            obj,
            CCAName,
            CCARank,
            choiceNumber,
            location,
            fitness,
            studentIndex,
            CCAIndex,
        ):
            """Add students to CCAs and add CCAs to students."""
            obj.StudentData[studentIndex]["CCA"] = CCAName
            obj.StudentData[studentIndex]["CCARank"] = CCARank
            obj.StudentData[studentIndex]["choiceNumbers"]["main"] = choiceNumber
            obj.StudentData[studentIndex]["locationOfAllocation"] = location
            obj.CCAData[CCAIndex]["selectedStudents"].append(
                {
                    "studentIndex": studentIndex,
                    "name": obj.StudentData[studentIndex]["name"],
                    "rank": CCARank,
                    "choice": choiceNumber,
                    "location": location,
                    "fitness": fitness,
                }
            )

        def DeleteStudent(obj, studentIndex, eliminated, CCAIndex):
            """Delete student from CCA and CCA from student."""
            obj.StudentData[studentIndex]["CCA"] = None
            obj.StudentData[studentIndex]["CCARank"] = None
            obj.StudentData[studentIndex]["choiceNumbers"]["main"] = 0
            obj.StudentData[studentIndex]["locationOfAllocation"] = None
            obj.CCAData[CCAIndex]["selectedStudents"].pop(eliminated)

        print("Allocating students.")
        logging.info("Allocation.Allocate: Allocating students.")
        # Popularity
        for student in self.StudentData:
            for choiceNumber in range(len(student["mainChoices"])):
                for CCA in range(len(self.CCAData)):
                    if (
                        student["mainChoices"][choiceNumber]
                        == self.CCAData[CCA]["name"]
                    ):
                        self.CCAData[CCA]["Popularity"] += (
                            self.Choices["main"] - choiceNumber
                        )
        # Cycle through all students until all students have been allocated (choice + shortlist)
        studentIndex = 0
        print("Allocating via selection and shortlist.")
        logging.info("Allocation.Allocate: Allocating via selection and shortlist.")
        allocated = 0  # Number of students allocated
        while allocated < len(
            self.StudentData
        ):  # While all students have not been allocated yet:
            # Reset student index
            if studentIndex == len(self.StudentData):
                studentIndex = 0
            # If the student has not been allocated
            # See whether they are selected by a CCA and has chosen that CCA,
            #   and if the CCA has available space,
            #       allocate
            #   otherwise
            #       see if another student could be reallocated
            #           Check if there is a student who is lower ranked by the CCA
            #               If found, remove old student and add new student
            #                   Old student would be given a new CCA with further rounds of selections
            for choiceNumber in range(
                len(self.StudentData[studentIndex]["mainChoices"])
            ):
                if self.StudentData[studentIndex]["CCA"] == None:
                    if (
                        self.StudentData[studentIndex]["mainChoices"][choiceNumber]
                        in self.StudentData[studentIndex]["shortlist"].keys()
                    ):
                        for CCA in range(len(self.CCAData)):
                            if (
                                self.StudentData[studentIndex]["mainChoices"][
                                    choiceNumber
                                ]
                                == self.CCAData[CCA]["name"]
                            ):
                                if (
                                    len(self.CCAData[CCA]["selectedStudents"])
                                    < self.CCAData[CCA]["allocated"]
                                ):
                                    AddStudent(
                                        self,
                                        self.CCAData[CCA]["name"],
                                        self.StudentData[studentIndex]["shortlist"][
                                            self.CCAData[CCA]["name"]
                                        ],
                                        choiceNumber + 1,
                                        0,
                                        self.StudentData[studentIndex]["fitnessScore"],
                                        studentIndex,
                                        CCA,
                                    )
                                else:  # Find which student to displace
                                    worstrank = 0  # Student with worst rank in CCA
                                    eliminated = (
                                        None
                                    )  # Index of which student gets displaced (self.StudentData)
                                    for selectedStudent in range(
                                        len(self.CCAData[CCA]["selectedStudents"])
                                    ):  # For all students
                                        if (
                                            self.CCAData[CCA]["selectedStudents"][
                                                selectedStudent
                                            ]["rank"]
                                            > worstrank
                                        ):  # If current worst rank is better than the rank found now
                                            worstrank = self.CCAData[CCA][
                                                "selectedStudents"
                                            ][selectedStudent][
                                                "rank"
                                            ]  # Change worst rank
                                            eliminated = (
                                                selectedStudent
                                            )  # Change eliminated index to worst ranked student in CCA
                                    if worstrank > self.CCAData[CCA]["shortlist"].index(
                                        self.StudentData[studentIndex]["name"]
                                    ):  # If worst rank student is worse than current student (self.StudentData[i])
                                        # Replace eliminated student
                                        DeleteStudent(
                                            self,
                                            self.CCAData[CCA]["selectedStudents"][
                                                eliminated
                                            ]["studentIndex"],
                                            eliminated,
                                            CCA,
                                        )
                                        AddStudent(
                                            self,
                                            self.CCAData[CCA]["name"],
                                            self.StudentData[studentIndex]["shortlist"][
                                                self.CCAData[CCA]["name"]
                                            ],
                                            choiceNumber + 1,
                                            0,
                                            self.StudentData[studentIndex][
                                                "fitnessScore"
                                            ],
                                            studentIndex,
                                            CCA,
                                        )
                                        allocated -= 1
            allocated += 1
            studentIndex += 1
        # Cycle through all remaining students (special)
        print("Allocating students via achievements.")
        logging.info("Allocation.Allocate: Achievements.")
        # Reset
        studentIndex = 0
        allocated = 0
        print("Allocating CCAs via achievements.")
        logging.info("Allocation.Allocate: Allocating CCAs via achievements.")
        while allocated < len(self.StudentData):
            if studentIndex == len(self.StudentData):
                studentIndex = 0
            # Same as above but finding students who have not been allocated and have art, music or special achievements
            for choiceNumber in range(
                len(self.StudentData[studentIndex]["mainChoices"])
            ):
                for CCA in range(len(self.CCAData)):
                    if self.StudentData[studentIndex]["CCA"] == None:
                        if (
                            self.StudentData[studentIndex]["mainChoices"][choiceNumber]
                            == self.CCAData[CCA]["name"]
                        ):
                            # Check whether CCA is type basic, art or music and matches certain criteria
                            if (
                                (
                                    self.CCAType[self.CCAData[CCA]["name"]] == "m"
                                    and self.StudentData[studentIndex]["musicInterest"]
                                    == True
                                )
                                or (
                                    self.CCAType[self.CCAData[CCA]["name"]] == "a"
                                    and self.StudentData[studentIndex]["artInterest"]
                                    == True
                                )
                                or self.StudentData[studentIndex]["specialCCA"]
                                == self.CCAData[CCA]["name"]
                            ):
                                # Check for vacancy
                                if (
                                    len(self.CCAData[CCA]["selectedStudents"])
                                    < self.CCAData[CCA]["allocated"]
                                ):
                                    # Find shortlist rank of student if there is anyway
                                    try:
                                        rank = self.StudentData[studentIndex][
                                            "shortlist"
                                        ][self.CCAData[CCA]["name"]]
                                    except KeyError:
                                        rank = None
                                    # Add student
                                    AddStudent(
                                        self,
                                        self.CCAData[CCA]["name"],
                                        rank,
                                        choiceNumber + 1,
                                        1,
                                        self.StudentData[studentIndex]["fitnessScore"],
                                        studentIndex,
                                        CCA,
                                    )
                                # If no vacancy, displace
                                else:
                                    # Find eliminatable student by choice
                                    worstchoice = 0
                                    eliminated = None
                                    for selectedStudent in range(
                                        len(self.CCAData[CCA]["selectedStudents"])
                                    ):
                                        # If selected student was allocated by achievements
                                        #   If choice is lower than worstchoice, make them the eliminated student
                                        #   However if same, randomise whether to replace eliminated student
                                        if (
                                            self.CCAData[CCA]["selectedStudents"][
                                                selectedStudent
                                            ]["location"]
                                            == 1
                                        ):
                                            if (
                                                self.CCAData[CCA]["selectedStudents"][
                                                    selectedStudent
                                                ]["choice"]
                                                > worstchoice
                                            ):
                                                worstchoice = self.CCAData[CCA][
                                                    "selectedStudents"
                                                ][selectedStudent]["choice"]
                                                eliminated = selectedStudent
                                            elif (
                                                self.CCAData[CCA]["selectedStudents"][
                                                    selectedStudent
                                                ]["choice"]
                                                == worstchoice
                                            ):
                                                if bool(random.randint(0, 1)):
                                                    eliminated = selectedStudent
                                    # If eliminatable student found and worst choice lower than current student, replace
                                    if eliminated == None:
                                        pass
                                    elif worstchoice > choiceNumber:
                                        try:
                                            rank = self.StudentData[studentIndex][
                                                "shortlist"
                                            ][self.CCAData[CCA]["name"]]
                                        except KeyError:
                                            rank = None
                                            DeleteStudent(
                                                self,
                                                self.CCAData[CCA]["selectedStudents"][
                                                    eliminated
                                                ]["studentIndex"],
                                                eliminated,
                                                CCA,
                                            )
                                        AddStudent(
                                            self,
                                            self.CCAData[CCA]["name"],
                                            rank,
                                            choiceNumber + 1,
                                            1,
                                            self.StudentData[studentIndex][
                                                "fitnessScore"
                                            ],
                                            studentIndex,
                                            CCA,
                                        )
                                        allocated -= 1
            allocated += 1
            studentIndex += 1
        # Allocate by pure CCA rank
        print("Allocating students only with shortlist ranks.")
        logging.info("Allocation.Allocate: Pure ranks.")
        studentIndex = 0
        allocated = 0
        while allocated < len(self.StudentData):
            if studentIndex == len(self.StudentData):
                studentIndex = 0
            for CCA in range(len(self.CCAData)):
                if self.StudentData[studentIndex]["CCA"] == None:
                    if (
                        self.StudentData[studentIndex]["name"]
                        in self.CCAData[CCA]["shortlist"]
                    ):
                        if (
                            len(self.CCAData[CCA]["selectedStudents"])
                            < self.CCAData[CCA]["allocated"]
                        ):
                            AddStudent(
                                self,
                                self.CCAData[CCA]["name"],
                                self.CCAData[CCA]["shortlist"].index(
                                    self.StudentData[studentIndex]["name"]
                                ),
                                0,
                                2,
                                self.StudentData[studentIndex]["fitnessScore"],
                                studentIndex,
                                CCA,
                            )
                        else:
                            worstrank = 0
                            eliminated = None
                            for selectedStudent in range(
                                len(self.CCAData[CCA]["selectedStudents"])
                            ):  # For all students
                                if (
                                    self.CCAData[CCA]["selectedStudents"][
                                        selectedStudent
                                    ]["rank"]
                                    > worstrank
                                    and self.CCAData[CCA]["selectedStudents"][
                                        selectedStudent
                                    ]["location"]
                                    == 2
                                ):  # If current worst rank is better than the rank found now
                                    worstrank = self.CCAData[CCA]["selectedStudents"][
                                        selectedStudent
                                    ][
                                        "rank"
                                    ]  # Change worst rank
                                    eliminated = (
                                        selectedStudent
                                    )  # Change eliminated index to worst ranked student in CCA
                            if worstrank > self.CCAData[CCA]["shortlist"].index(
                                self.StudentData[studentIndex]["name"]
                            ):  # If worst rank student is worse than current student (self.StudentData[i])
                                # Replace eliminated student
                                DeleteStudent(
                                    self,
                                    self.CCAData[CCA]["selectedStudents"][eliminated][
                                        "studentIndex"
                                    ],
                                    eliminated,
                                    CCA,
                                )
                                AddStudent(
                                    self,
                                    self.CCAData[CCA]["name"],
                                    self.StudentData[studentIndex]["shortlist"][
                                        self.CCAData[CCA]["name"]
                                    ],
                                    0,
                                    2,
                                    self.StudentData[studentIndex]["fitnessScore"],
                                    studentIndex,
                                    CCA,
                                )
                                allocated -= 1
            allocated += 1
        # Allocate for sports with healthStats using each student's fitnessScore
        print("Allocating students via fitness of students.")
        logging.info("Allocation.Allocate: Fitness")
        studentIndex = 0
        allocated = 0
        if self.SheetOrder["healthStats"] != None:
            while allocated < len(self.StudentData):
                if studentIndex == len(self.StudentData):
                    studentIndex = 0
                for choiceNumber in range(
                    len(self.StudentData[studentIndex]["mainChoices"])
                ):
                    for CCA in range(len(self.CCAData)):
                        if self.StudentData[studentIndex]["CCA"] == None:
                            if self.CCAType[self.CCAData[CCA]["name"]] == "s":
                                if (
                                    len(self.CCAData[CCA]["selectedStudents"])
                                    < self.CCAData[CCA]["allocated"]
                                ):
                                    try:
                                        rank = self.StudentData[studentIndex][
                                            "shortlist"
                                        ][self.CCAData[CCA]["name"]]
                                    except KeyError:
                                        rank = None
                                    AddStudent(
                                        self,
                                        self.CCAData[CCA]["name"],
                                        rank,
                                        choiceNumber + 1,
                                        3,
                                        self.StudentData[studentIndex]["fitnessScore"],
                                        studentIndex,
                                        CCA,
                                    )
                                else:
                                    worsthealth = (
                                        256
                                    )  # Arbitrary value, student can't possibly be that unhealthy
                                    eliminated = 0
                                    for selectedStudent in range(
                                        len(self.CCAData[CCA]["selectedStudents"])
                                    ):
                                        if (
                                            self.CCAData[CCA]["selectedStudents"][
                                                selectedStudent
                                            ]["location"]
                                            == 3
                                        ):
                                            if (
                                                self.CCAData[CCA]["selectedStudents"][
                                                    selectedStudent
                                                ]["fitness"]
                                                < worsthealth
                                            ):
                                                worsthealth = self.CCAData[CCA][
                                                    "selectedStudents"
                                                ][selectedStudent]["fitness"]
                                                eliminated = selectedStudent
                                            elif (
                                                self.CCAData[CCA]["selectedStudents"][
                                                    selectedStudent
                                                ]["fitness"]
                                                == worsthealth
                                            ):
                                                if bool(random.randint(0, 1)):
                                                    eliminated = selectedStudent
                                    if (
                                        worsthealth
                                        < self.StudentData[studentIndex]["fitnessScore"]
                                    ):
                                        try:
                                            rank = self.StudentData[studentIndex][
                                                "shortlist"
                                            ][self.CCAData[CCA]["name"]]
                                        except KeyError:
                                            rank = None
                                        DeleteStudent(
                                            self,
                                            self.CCAData[CCA]["selectedStudents"][
                                                eliminated
                                            ]["studentIndex"],
                                            eliminated,
                                            CCA,
                                        )
                                        AddStudent(
                                            self,
                                            self.CCAData[CCA]["name"],
                                            rank,
                                            choiceNumber + 1,
                                            3,
                                            self.StudentData[studentIndex][
                                                "fitnessScore"
                                            ],
                                            studentIndex,
                                            CCA,
                                        )
                                        allocated -= 1
                studentIndex += 1
                allocated += 1
        # Allocate by randomisation
        print("Allocating students randomly.")
        logging.info("Allocating.Allocation: Random")
        studentIndex = 0
        allocated = 0
        while allocated < len(self.StudentData):
            if studentIndex == len(self.StudentData):
                studentIndex = 0
            for CCA in range(len(self.CCAData)):
                if self.StudentData[studentIndex]["CCA"] != None:
                    pass
                else:
                    if (
                        len(self.CCAData[CCA]["selectedStudents"])
                        < self.CCAData[CCA]["allocated"]
                    ):
                        try:
                            rank = self.StudentData[studentIndex]["shortlist"][
                                self.CCAData[CCA]["name"]
                            ]
                        except KeyError:
                            rank = None
                        try:
                            choiceNumber = self.StudentData[studentIndex][
                                "mainChoices"
                            ].index(self.CCAData[CCA]["name"])
                        except ValueError:
                            choiceNumber = 0
                        AddStudent(
                            self,
                            self.CCAData[CCA]["name"],
                            rank,
                            choiceNumber,
                            4,
                            self.StudentData[studentIndex]["fitnessScore"],
                            studentIndex,
                            CCA,
                        )
            allocated += 1
            studentIndex += 1
        # for i in self.StudentData:
        # print(f'Index: {i["no"]} Name: {i["name"]} CCA: {i["CCA"]} Choice: {i["choiceNumbers"]["main"]} Rank: {i["CCARank"]} Fitness level: {i["fitnessScore"]}')
        print("Allocation complete.")
        logging.info("Allocation.Allocate: Allocation complete.")
        self.Allocated = True

    def SaveToFile(self):
        """
        Saves allocation to output file. Not compulsory but why wouldn't you?

        Raises
        ------
        PermissionError
            If the input workbook is open in another program.
        UnknownError
            If an unknown error occurs
        """
        print("Saving to file:", self.Path + self.FileName)
        logging.info("Allocation.SaveToFile: Saving to file.")
        # For each student, find their row in studentList and add in the data from the Allocation.Allocate
        for studentInfo in self.StudentData:
            # Retrieve cell object for CCA and set value to student's CCA
            cell = self.Worksheets[self.SheetOrder["studentList"]].cell(
                row=(studentInfo["no"] + 1), column=7
            )
            cell.value = studentInfo["CCA"]
            # Choice of the student (If they did not choose CCA, the cell will be left blank (cell.value = None))
            cell = self.Worksheets[self.SheetOrder["studentList"]].cell(
                row=(studentInfo["no"] + 1), column=8
            )
            if studentInfo["choiceNumbers"]["main"] != 0:
                cell.value = studentInfo["choiceNumbers"]["main"]
            else:
                cell.value = None
            # Shortlist ranking in the CCA
            cell = self.Worksheets[self.SheetOrder["studentList"]].cell(
                row=(studentInfo["no"] + 1), column=9
            )
            if studentInfo["CCARank"] != None:
                cell.value = studentInfo["CCARank"] + 1
            else:
                cell.value = None
        os.chdir(self.Path)  # Go to location of workbook
        # Save the file
        try:
            self.Workbook.save(filename=self.FileName)
        except PermissionError as e:
            logging.error("Allocation.SaveToFile: Could not save file.", exc_info=True)
            raise UnableToSaveFileError(
                "Could not save to file. Make sure that input file is not opened and has not been modified or deleted before the program has finished executing."
            )
        except UnknownError as e:
            logging.critical(
                "Allocation.SaveToFile: Unknown error when saving file.", exc_info=True
            )
            raise UnknownError(e)
        print("Data saved to file.")
        logging.info("Allocation.SaveToFile: Data saved to file.")

    def Lottery(self, **kwargs):
        """
        Returns/prints of tuple or string students data include CCA allocated.

        **kwargs Parameters
        -------------------
        print: bool
            Whether to print or return values.
        studentIndex: int or list of int
            Which student to get values, may be used in conjunction with 'class'. If left empty or a value other than integer is passed through, it will cycle through the students.
        class: str or list of str
            Which class of students to cycle through. If used with 'studentIndex', the method will get the student indicated in 'studentIndex' in 'class'.

        Raises
        ------
        no
        """
        print("Beginning lottery.")
        logging.info("Allocation.Lottery: Setting up keys.")
        options = {"print": False, "studentIndex": None, "class": None}
        # Check if print is there, if not found, assume print == false and return data
        if "_print" in kwargs.keys():
            if type(kwargs["_print"]) != bool:
                logging.warning(
                    "Allocation.Lottery: _print in kwargs not a boolean. Print will be assumed to be false (return)."
                )
            else:
                options["print"] = kwargs["_print"]
        else:
            logging.warning("Allocation.Lottery: _print not in kwargs.")
        # See if user specified any students to get data about. If there is, input everything into a list
        if "_studentIndex" in kwargs.keys():
            if isinstance(kwargs["_studentIndex"], int):
                options["studentIndex"] = [kwargs["_studentIndex"]]
            elif type(kwargs["_studentIndex"]) == list:
                options["studentIndex"] = kwargs["_studentIndex"]
            elif kwargs["_studentIndex"] == None:
                options["studentIndex"] = None
        # Class of students
        if "_class" in kwargs.keys():
            if type(kwargs["_class"]) == str:
                options["class"] = [kwargs["_class"]]
            elif type(kwargs["_class"]) == list:
                options["class"] = kwargs["_class"]
            elif kwargs["_class"] == None:
                options["class"] = None
        print(options["print"], options["studentIndex"], options["class"])
        # Bunch of variables to make data mining easier
        returnTuple = []  # List first for easier appending
        classList = []  # List for class
        returnString = ""
        text = "No.: {0} Name: {1} CCA: {2} Ranking by CCA: {3} Choice: {4} Fitness: {5} How student was allocated: {6}"
        allocationMethod = [
            "Choice and wanted by CCA",
            "Achievements",
            "Wanted by CCA (but not necessarily because choice)",
            "Fitness Level",
            "Randomisation",
        ]
        # If no class specified
        #   If no students specified
        #       print or return all data (porad)
        #   If students specified
        #       Cycle through all students until student found
        #           porad
        # If class specified
        #   If no students specified
        #       Find all the students in the classes
        #           porad
        #   If students specified
        #       Find all students and put them in class list
        #           Locate student index using provided student index
        #               porad
        if options["class"] == None:
            if options["studentIndex"] == None:
                for student in self.StudentData:
                    if options["print"]:
                        print(
                            text.format(
                                student["no"],
                                student["name"],
                                student["CCA"],
                                student["CCARank"],
                                student["choiceNumbers"]["main"],
                                student["fitnessScore"],
                                allocationMethod[student["locationOfAllocation"]],
                            )
                        )
                    else:
                        returnTuple.append(
                            (
                                student["no"],
                                student["name"],
                                student["CCA"],
                                student["CCARank"],
                                student["choiceNumbers"]["main"],
                                student["fitnessScore"],
                                allocationMethod[student["locationOfAllocation"]],
                            )
                        )
            else:
                for selectedStudent in options["studentIndex"]:
                    for student in self.StudentData:
                        if selectedStudent == student["no"]:
                            if options["print"]:
                                print(
                                    text.format(
                                        student["no"],
                                        student["name"],
                                        student["CCA"],
                                        student["CCARank"],
                                        student["choiceNumbers"]["main"],
                                        student["fitnessScore"],
                                        allocationMethod[
                                            student["locationOfAllocation"]
                                        ],
                                    )
                                )
                            else:
                                returnTuple.append(
                                    (
                                        student["no"],
                                        student["name"],
                                        student["CCA"],
                                        student["CCARank"],
                                        student["choiceNumbers"]["main"],
                                        student["fitnessScore"],
                                        allocationMethod[
                                            student["locationOfAllocation"]
                                        ],
                                    )
                                )
        else:
            for className in options["class"]:
                returnTuple.append([])
                if options["studentIndex"] == None:
                    studentIndexNumber = 1
                    for student in self.StudentData:
                        if student["class"] == className:
                            if options["print"]:
                                print(
                                    text.format(
                                        studentIndexNumber,
                                        student["name"],
                                        student["CCA"],
                                        student["CCARank"],
                                        student["choiceNumbers"]["main"],
                                        student["fitnessScore"],
                                        allocationMethod[
                                            student["locationOfAllocation"]
                                        ],
                                    )
                                )
                            else:
                                returnTuple[-1].append(
                                    (
                                        studentIndexNumber,
                                        student["name"],
                                        student["CCA"],
                                        student["CCARank"],
                                        student["choiceNumbers"]["main"],
                                        student["fitnessScore"],
                                        allocationMethod[
                                            student["locationOfAllocation"]
                                        ],
                                    )
                                )
                        studentIndexNumber += 1
                else:
                    for student in self.StudentData:
                        if student["class"] == className:
                            classList.append(student)
                    for studentIndex in range(len(classList)):
                        for selectedStudent in options["studentIndex"]:
                            if studentIndex == selectedStudent - 1:
                                student = classList[studentIndex]
                                if options["print"]:
                                    print(
                                        text.format(
                                            selectedStudent,
                                            student["name"],
                                            student["CCA"],
                                            student["CCARank"],
                                            student["choiceNumbers"]["main"],
                                            student["fitnessScore"],
                                            allocationMethod[
                                                student["locationOfAllocation"]
                                            ],
                                        )
                                    )
                                else:
                                    returnTuple[-1].append(
                                        (
                                            selectedStudent,
                                            student["name"],
                                            student["CCA"],
                                            student["CCARank"],
                                            student["choiceNumbers"]["main"],
                                            student["fitnessScore"],
                                            allocationMethod[
                                                student["locationOfAllocation"]
                                            ],
                                        )
                                    )
                returnTuple[-1] = tuple(returnTuple[-1])  # Compile list into tuple
        logging.info("Allocation.Lottery: Info collected.")
        # If return data
        #   Return tuple
        if options["print"] == False:
            if returnTuple == ():
                return None
            else:
                return tuple(returnTuple)


class Template:
    """
    Simple creation of a template
    """

    def __init__(self, **kwargs):
        """
        Creates a template input workbook.

        **kwargs Parameters
        -------------------
        path: str
            Path to where template should be
        fileName: str
            Desired file name of template
        CCAs: list
            What CCAs to make shortlist sheets for
        sheetOrder: dict
            Where the admin sheets should go
        choices: dict
            How many choices each student gets. Configures "choices" admin sheet

        Raises
        ------
        no
        """
        print("Creating template workbook.")
        logging.info("Template.__init__: Creating template workbook.")
        self.Workbook = openpyxl.Workbook()
        if "path" in kwargs.keys():
            self.Path = pathlib.Path(kwargs["path"])
        else:
            self.Path = pathlib.Path(os.getcwd())
        if "fileName" in kwargs.keys():
            self.FileName = kwargs["fileName"]
        else:
            self.FileName = "ccallocator-template.xlsx"
        self.FilePath = pathlib.Path(self.Path, self.FileName)
        if "CCAs" in kwargs.keys():
            self.CCAList = kwargs["CCAs"]
        else:
            self.CCAList = []
        self.SheetOrder = {
            "studentList": False,
            "healthStats": False,
            "music": False,
            "art": False,
            "special": False,
            "CCAList": False,
            "choices": False,
        }
        if "sheetOrder" in kwargs.keys():
            for sheet, value in kwargs["sheetOrder"].items():
                if sheet in self.SheetOrder.keys() and value:
                    self.SheetOrder[sheet] = True
        self.Choices = {"main": 1, "other": None}
        if "choices" in kwargs.keys():
            for choiceType, value in kwargs["choices"].items():
                if choiceType in self.Choices.keys() and (
                    isinstance(value, int) or value == None
                ):
                    self.Choices[choiceType] = value
        self.Worksheet = {}
        for sheet in self.SheetOrder.keys():
            if self.SheetOrder[sheet]:
                self.Worksheet[sheet] = self.Workbook.create_sheet(sheet)
        for sheet in self.Worksheet:
            if sheet == "studentList":
                self.Worksheet[sheet]["A1"] = "SN"
                self.Worksheet[sheet]["B1"] = "NO"
                self.Worksheet[sheet]["C1"] = "NAME"
                self.Worksheet[sheet]["D1"] = "IDNO"
                self.Worksheet[sheet]["E1"] = "CLASS"
                self.Worksheet[sheet]["F1"] = "DATE OF BIRTH"
                self.Worksheet[sheet]["G1"] = "ALLOCATED CCA"
                self.Worksheet[sheet]["H1"] = "CHOICE #"
                self.Worksheet[sheet]["I1"] = "RANK"
            elif sheet == "healthStats":
                self.Worksheet[sheet]["A1"] = "NO"
                self.Worksheet[sheet]["B1"] = "NAME"
                self.Worksheet[sheet]["C1"] = "IDNO"
                self.Worksheet[sheet]["D1"] = "AGE"
                self.Worksheet[sheet]["E1"] = "CLASS"
            elif sheet == "music" or sheet == "art":
                self.Worksheet[sheet]["A1"] = "IDNO"
                self.Worksheet[sheet]["B1"] = "CLASS"
                self.Worksheet[sheet]["C1"] = "REMARKS"
            elif sheet == "special":
                self.Worksheet[sheet]["A1"] = "NAME"
                self.Worksheet[sheet]["B1"] = "DOMAIN"
                self.Worksheet[sheet]["C1"] = "CCA"
                self.Worksheet[sheet]["D1"] = "CLASS"
            elif sheet == "CCAList":
                self.Worksheet[sheet]["A1"] = "NO"
                self.Worksheet[sheet]["B1"] = "CCA"
                self.Worksheet[sheet]["C1"] = "CCA LONG NAME"
                self.Worksheet[sheet]["D1"] = "ALLOCATED"
            elif sheet == "choices":
                self.Worksheet[sheet]["A1"] = "SN"
                self.Worksheet[sheet]["B1"] = "IDNO"
                self.Worksheet[sheet]["C1"] = "DATE OF BIRTH"
                self.Worksheet[sheet]["D1"] = "CLASS"
                self.Worksheet[sheet]["E1"] = "PAST CCA"
                if self.Choices["main"] == None:
                    self.Choices["main"] = 1
                for choice in range(1, self.Choices["main"] + 1):
                    cell = self.Worksheet[sheet].cell(row=1, column=5 + choice)
                    cell.value = "MAIN CHOICE" + str(choice + 1)
                if self.Choices["other"] != None:
                    for choice in range(self.Choices["other"]):
                        cell = self.Worksheet[sheet].cell(
                            row=1, column=5 + self.Choices["main"] + choice
                        )
                        cell.value = "OTHER CHOICE" + str(choice + 1)
        tempobj = self.Workbook.get_sheet_by_name("Sheet")
        self.Workbook.remove_sheet(tempobj)
        for CCA in self.CCAList:
            self.Worksheet[CCA] = self.Workbook.create_sheet(CCA)
            self.Worksheet[CCA]["A1"] = "RANK"
            self.Worksheet[CCA]["B1"] = "NAME"
            self.Worksheet[CCA]["C1"] = "CLASS"
        self.Workbook.save(self.FilePath)
        print("Template workbook created.")
        logging.info("Template.__init__: Template workbook created.")
